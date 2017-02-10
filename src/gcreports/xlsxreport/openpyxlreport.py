import os

import pandas
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from gcreports.xlsxreport.excelreport import ExcelReport
from gcreports.xlsxreport.mynamedstyles import MyNamedStyles

class OpenpyxlReport(ExcelReport):
    """
    Reports from dataframe to Excel based on openpyxl
    Недработано
    but
    Note: Support for OpenPyxl v2.2 is currently EXPERIMENTAL (GH7565).
    """

    # default_dir_reports = 'U:/tables'

    def __init__(self, filename, sheet='Sheet1', template_file=None):
        super(OpenpyxlReport, self).__init__(filename=filename, sheet=sheet, engine='openpyxl')
        self.styles = MyNamedStyles(template_file)
        # self.filename = filename
        # Create a Pandas Excel writer using XlsxWriter as the engine.
        # writer = pandas.ExcelWriter(filename, engine='xlsxwriter', datetime_format=datetime_format)
        # self.writer = pandas.ExcelWriter(filename, engine='openpyxl')

        # self.workbook = self.writer.book
        # self.worksheet = self.workbook.active
        # self.worksheet.title = sheet

    # def save(self):
    #     self.workbook.save(self.filename)

    def _rows_to_sheet(self, rows, name, start_row=2, start_col=1, header=False, index=True):
        head = True
        col_head = True
        frmt_header = self.styles.get_style(name)
        # Заголовок
        if frmt_header:
            ce = self.worksheet.cell(row=start_row-1, column=start_col, value=frmt_header.value)
            frmt_header.set_on_cell(ce)

        for r_idx, row in enumerate(rows, start_row):
            col_head = True
            for c_idx, value in enumerate(row, start_col):
                ce = self.worksheet.cell(row=r_idx, column=c_idx, value=value)
                if head and header:
                    self.styles.set_style(ce, name)
                col_head = False
            head = False


    def add_df_test(self, datafarame):
        for r in dataframe_to_rows(datafarame, index=True, header=True):
            # if type(r) == tuple:
            #     r =
            self.worksheet.append(r)

    def add_dataframe(self, dataframe, name, start_row=2, start_col=1, header=False, index=True):

        rows = dataframe_to_rows(dataframe, index=index, header=header)
        head = True
        col_head = True
        frmt_header = self.styles.get_style(name)
        # Заголовок
        if frmt_header:
            ce = self.worksheet.cell(row=start_row - 1, column=start_col, value=frmt_header.value)
            frmt_header.set_on_cell(ce)

        for r_idx, row in enumerate(rows, start_row):
            col_head = True
            for c_idx, value in enumerate(row, start_col):
                if value:
                    ce = self.worksheet.cell(row=r_idx, column=c_idx, value=value)
                    # ce = self.worksheet.cell(row=r_idx, column=c_idx, value=value)
                    if head and header:
                        self.styles.set_style(ce, name)
                col_head = False
            head = False